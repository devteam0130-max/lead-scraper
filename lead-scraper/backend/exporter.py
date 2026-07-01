import io
import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Cores
COR_HEADER_BG = "1e3a5f"   # Azul escuro
COR_HEADER_FG = "FFFFFF"   # Branco
COR_LINHA_ALT = "f0f4f8"   # Azul muito claro
COR_LINHA_PAR = "FFFFFF"   # Branco
COR_WPP_TEXTO = "155724"   # Verde escuro
COR_WPP_BG = "d4edda"      # Verde claro
COR_LINK = "1a56b0"        # Azul link

# Colunas da planilha (nesta ordem)
COLUNAS = [
    "#",
    "Nome",
    "Telefone",
    "WhatsApp",
    "Endereço",
    "Cidade",
    "Site",
    "Avaliação",
    "Nº Avaliações",
    "Categoria",
]

# Larguras aproximadas de cada coluna (em caracteres)
COL_WIDTHS = [5, 38, 18, 22, 45, 22, 35, 11, 14, 28]


def generate_xlsx(resultados: list[dict]) -> bytes:
    """
    Gera o arquivo .xlsx com os resultados de leads.
    Retorna os bytes prontos para download.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Leads"

    # --- Cabeçalho ---
    _escrever_cabecalho(ws)

    # --- Linhas de dados ---
    for idx, r in enumerate(resultados, start=1):
        _escrever_linha(ws, idx + 1, idx, r)

    # --- Freeze da primeira linha ---
    ws.freeze_panes = "A2"

    # --- Largura das colunas ---
    for i, largura in enumerate(COL_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = largura

    # --- Altura padrão das linhas ---
    ws.row_dimensions[1].height = 22
    for row_idx in range(2, len(resultados) + 2):
        ws.row_dimensions[row_idx].height = 18

    # --- Salvar em memória e retornar bytes ---
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _escrever_cabecalho(ws):
    """Escreve a linha de cabeçalho com fundo azul escuro e texto branco em negrito."""
    fill_header = PatternFill("solid", start_color=COR_HEADER_BG)
    font_header = Font(bold=True, color=COR_HEADER_FG, name="Arial", size=10)
    align_center = Alignment(horizontal="center", vertical="center")

    for col_idx, nome_col in enumerate(COLUNAS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=nome_col)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center


def _escrever_linha(ws, row_idx: int, num: int, r: dict):
    """Escreve uma linha de dados com formatação alternada e destaque especial para WhatsApp."""
    # Cor de fundo alternada
    cor_bg = COR_LINHA_ALT if num % 2 == 0 else COR_LINHA_PAR
    fill_normal = PatternFill("solid", start_color=cor_bg)
    font_normal = Font(name="Arial", size=9)
    align_v = Alignment(vertical="center", wrap_text=False)

    # Extrair cidade do endereço
    endereco = r.get("endereco") or ""
    cidade = _extrair_cidade(endereco)

    # Dados da linha na ordem das colunas
    linha_dados = [
        num,
        r.get("nome") or "",
        r.get("telefone") or "",
        None,                           # WhatsApp — tratado separadamente
        endereco,
        cidade,
        r.get("site") or "",
        r.get("avaliacao") or "",
        r.get("num_avaliacoes") or "",
        r.get("categoria") or "",
    ]

    for col_idx, valor in enumerate(linha_dados, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=valor)
        cell.font = font_normal
        cell.fill = fill_normal
        cell.alignment = align_v

    # --- Coluna WhatsApp (índice 4) com formatação especial ---
    wpp = r.get("whatsapp")
    wpp_cell = ws.cell(row=row_idx, column=4)

    if wpp:
        wpp_cell.value = wpp
        wpp_cell.font = Font(name="Arial", size=9, color=COR_WPP_TEXTO, bold=True)
        wpp_cell.fill = PatternFill("solid", start_color=COR_WPP_BG)
        wpp_cell.alignment = align_v
        # Link clicável para abrir o WhatsApp
        if wpp.startswith("http"):
            wpp_cell.hyperlink = wpp
        else:
            wpp_cell.hyperlink = f"https://wa.me/{wpp}"
    else:
        wpp_cell.value = ""
        wpp_cell.font = font_normal
        wpp_cell.fill = fill_normal
        wpp_cell.alignment = align_v

    # --- Coluna Site (índice 7) com hyperlink ---
    site = r.get("site") or ""
    if site:
        site_cell = ws.cell(row=row_idx, column=7)
        site_cell.hyperlink = site if site.startswith("http") else f"https://{site}"
        site_cell.font = Font(name="Arial", size=9, color=COR_LINK, underline="single")


def _extrair_cidade(endereco: str) -> str:
    """
    Tenta extrair o nome da cidade de um endereço brasileiro.
    Ex: "Rua XYZ, 123 - Bairro, Salvador - BA, 40000-000" → "Salvador"
    """
    if not endereco:
        return ""

    # Tentar padrão "Cidade - UF"
    match = re.search(r",\s*([^,]+?)\s*-\s*[A-Z]{2}\b", endereco)
    if match:
        return match.group(1).strip()

    # Fallback: penúltima parte separada por vírgula
    partes = [p.strip() for p in endereco.split(",") if p.strip()]
    if len(partes) >= 2:
        # Remover CEP e sigla de estado da última parte
        candidato = partes[-2]
        candidato = re.sub(r"\s*-\s*[A-Z]{2}$", "", candidato).strip()
        return candidato

    return ""
